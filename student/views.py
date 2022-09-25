import datetime
from django.http import Http404, HttpResponse, HttpResponseNotFound
from django.shortcuts import redirect, render
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from requests import session
from django.contrib.auth.decorators import login_required
from .models import teacher_profile

from school.models import School
from student.models import *
import openpyxl
from datetime import date
from django.utils.crypto import get_random_string
from django.contrib import messages


# Create your views here.



# profile of the student
# Ayon
@login_required(login_url='/login')
def profile_student(request):
    # gets the teacher id of the teacher profile and checks if it exists
    if request.user.groups.filter(name='school').exists():
        school = request.user.schooluserprofile
        schoolNavHeader = True
    else:
        school = request.user.teacherprofile.school
        schoolNavHeader = False

    if request.GET['student_id'] != '' and student_profile.objects.filter(id=request.GET['student_id'], school=school).exists():
        context = {
            'student' : student_profile.objects.get(id=request.GET['student_id']),
            'schoolNavHeader' : schoolNavHeader
        }
        return render(request, 'student/profile.html', context)
    else:
        return HttpResponseNotFound('<h2 style="text-align:center;margin-top:400px;">Student not Found(HTTP_404)<h2>')


# get all the student
# Ayon
@login_required(login_url='/school/login')
def students(request):
    if request.user.groups.filter(name='school').exists() or request.user.groups.filter(name='Teachers').exists():
        if request.user.groups.filter(name='school').exists():
            school = request.user.schooluserprofile
            schoolNavHeader = True
        else:
            school = request.user.teacherprofile.school
            schoolNavHeader = False

        students = student_profile.objects.filter(school = school)
        return render(request, 'student/students.html' , {'student' : students, 'schoolNavHeader' : schoolNavHeader})
    return HttpResponse("Working")


# login the user and set his usertype
# Ayon
def login_student(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            # User is authenticated

            # checks the user is Studentsor not
            if user.groups.filter(name='Students').exists():
                # if yes set the user type as Students and login

                # log in the teacher
                login(request, user)

                # starts a session teacher and and userType of the session is Teacher
                request.session['student'] = {
                'isLoggedIn' : True,
                'userType' : "Students"
                }
                return redirect('/student?msg=login success')
            else :
                return redirect('/student/login?msg=invalid group')
        else:
            return redirect("/student/login?msg=invalid credentials")
    return render(request, 'student/login.html')


# update the student_profile
# Ayon
@login_required(login_url='/teacher/login')
def edit_student(request):
    if request.user.is_authenticated:
        try:

            if request.method == 'POST':
                user = User.objects.get(id = request.user.id)
                user.first_name = request.POST['first_name']
                user.last_name = request.POST['last_name']
                user.save()

                if len(request.FILES) != 0 :
                    student_profile.objects.filter(user=user).update(img = request.FILES.get('img'))

                student = student_profile.objects.filter(user=user).update(
                    full_name  = f"{request.POST['first_name'] } {request.POST['last_name'] }",
                    phone = request.POST['phone'],
                    gender = request.POST['gender'],
                )
                
                student = student_profile.objects.get(user=user)
                if request.POST['father_name'] != '':
                    student.father_name = request.POST['father_name']

                if request.POST['mother_name'] != '':
                    student.mother_name = request.POST['mother_name']

                if request.POST['father_phone'] != '':
                    student.father_phone = request.POST['father_phone']

                if request.POST['mother_phone'] != '':
                    student.mother_phone = request.POST['mother_phone']

                if request.POST['roll_no'] != '':
                    student.roll_no = request.POST['roll_no']

                if request.POST['address'] != '':
                    student.address = request.POST['address']

                if request.POST['zipcode'] != '':
                    student.zipcode = request.POST['zipcode']     

                if len( request.FILES) != 0:
                    student.img = request.FILES['img']       

                if request.POST['dob'] != '':
                    student.dob = request.POST['dob']


                student.save()
                messages.success(request, f"Profile updated")
                return redirect('/student/update?msg=update success')
            else:
                user = User.objects.get(id = request.user.id)
                student = student_profile.objects.get(user=user)
                context = {
                    'student' : student,
                    'school_id' : student.school.id
                }
                return render(request, 'student/edit-student.html', context)
        except:
            messages.error(request, f"error while updatating")
            return redirect('/student/update/')        
    return redirect('/')


# logout for the student
# Ayon
def logout_student(request):
    try :
        del request.session['student']
    except:
        pass
    logout(request)
    return redirect('/')


# upload excel for the student
# Ayon
@login_required(login_url='/teacher/login')
def upload_student(request):
    if request.user.groups.filter(name='school').exists() or request.user.groups.filter(name='Teachers').exists():
        if request.user.groups.filter(name='school').exists():
            school = request.user.schooluserprofile
            schoolNavHeader = True
        else:
            school = request.user.teacherprofile.school
            schoolNavHeader = False

        try:

            if request.method == 'POST':

                if request.FILES['student_file']  != '':
                    data = openpyxl.load_workbook(request.FILES['student_file'])

                    data = data.active

                    data_list = []
                    count = 0
                    for row in range(2, data.max_row):
                        for col in data.iter_cols(1, 6): #for col in data.iter_cols(1, data.max_column):
                            if (str(col[row].value) != 'None' and  col[row].value != '' and col[row].value != None ):
                                    data_list.append(str(col[row].value))
                                    print(str(col[row].value))
                            else:
                                
                                messages.error(request, f'Students Upload Failed (Check : cell no [{col[row]}] #EMPTY_{str(col[row].value)})')
                                return redirect(f"/student/add/upload/?msg=upload failed (Check : cell no [{col[row]}] #template error)")

                        first_name = data_list[0]
                        last_name =  data_list[1]
                        
                        #GENDER FORMATTER
                        if(data_list[2].lower() == 'f'):
                            gender = "Female"
                        elif(data_list[2].lower() == 'm'):
                            gender = "Male"
                        elif(data_list[2].lower() == 'o'):
                            gender = "Others"
                        else:
                            messages.error(request, f'Students Upload Failed (Check : cell no [{row}, 3] GenderFormatError#Should be M/F/O : {data_list[2]})')
                            return redirect(f"/student/add/upload/?msg=upload failed (Check : cell no [{row}, 3] #GenderFormatError)") 

                        email = data_list[3]
                        phone = data_list[4]

                        if Classroom.objects.filter(classroom_id=data_list[5], school_id=school).exists():
                            classroom = Classroom.objects.get(classroom_id=data_list[5], school_id=school)
                        else:
                            messages.error(request, f"Students Upload Failed (Check : cell no [{row}, 6] ClassroomCodeError#Classroom can't be assigned: {data_list[5]})")
                            return redirect(f"/student/add/upload/?msg=upload failed (Check : cell no [{row}, 6] #ClassroomCodeError)") 

                        username = first_name+'_'+last_name
                        while User.objects.filter(username=username).exists():
                            username = first_name+'_'+last_name+'_'+str(get_random_string(4))
                        password = get_random_string(8)

            
                        full_name  = f"{first_name} {last_name}"

                        data_list.clear()

                        student = User.objects.create_user(username, email, password)
                        student.first_name = first_name
                        student.last_name = last_name
                        student.save()

                        # add the teacher to the Students group
                        Group.objects.get(name='Students').user_set.add(student)

                        # create the student profile for the student account
                        add_student_profile = student_profile.objects.create(
                            user = student, # the student account
                            school = school,
                            full_name  = full_name,
                            phone = phone,
                            gender = gender,
                            classroom = classroom,
                            psw = password
                        )
                        count = count + 1 
                    return redirect(f"/student/list?msg=upload-success-{count}")
                return redirect('/student/add/upload?id=')
            context = { 'schoolNavHeader' : schoolNavHeader }
            return render(request, 'student/add-upload.html', context)
        except:
            messages.error(request, 'Student Upload Failed')
            return redirect('/student/add/upload/')  
    return redirect('/school/school-login')


# student dash
# Ayon
@login_required(login_url='/student/login')
def dash(request):
    if request.user.is_authenticated and student_profile.objects.filter(user=User.objects.get(id=request.user.id)).exists():
        student = student_profile.objects.get(id=request.user.studentprofile.id)
        context = {
            'student' : student,
            'school_id' : student.school.id
        }
        return render(request, 'student/index.html', context)
    return redirect('/student/login')


# create the student account and student profile
# Ayon
@login_required(login_url='/school/school-login')
def create_student(request):
    # check if the school/teacher is logged in or not
    if request.user.groups.filter(name='school').exists() or request.user.groups.filter(name='Teachers').exists():
        if request.user.groups.filter(name='school').exists():
            school = request.user.schooluserprofile
            schoolNavHeader = True
        else:
            school = request.user.teacherprofile.school
            schoolNavHeader = False

        try:
            if request.method == 'POST':
                first_name = request.POST['first_name']
                last_name = request.POST['last_name']
                email = request.POST['email']

                phone = request.POST['phone']
                gender = request.POST['gender']
                classroom = request.POST['classroom']

                username = first_name+'_'+last_name
                while User.objects.filter(username=username).exists():
                    username = first_name+'_'+last_name+'_'+get_random_string(4)

                # PASSWORD
                password = get_random_string(8)

                # create a user account for the student
                student = User.objects.create_user(username, email, password)
                student.first_name = first_name
                student.last_name = last_name
                student.save()

                # add the student to the Students group
                Group.objects.get(name='Students').user_set.add(student)


                # create the student profile for the student account
                add_student_profile = student_profile.objects.create(
                    user = student, # the teacher account
                    school = school,
                    full_name  = f"{first_name} {last_name}",
                    phone = phone,
                    gender = gender,
                    classroom = Classroom.objects.get(id=int(classroom), school_id=school),
                    psw = password
                )
                messages.success(request, 'Student Added')
                return redirect(f'/student/profile?student_id={add_student_profile.id}&msg=add success')
            else:
                context = {
                    'schoolNavHeader' : schoolNavHeader,
                    'classroom' : Classroom.objects.filter(school_id=school)
                }
                return render(request, 'student/add-student.html', context)
        except:
            messages.error(request, 'Student Add Failed')
            return redirect('/student/add/')
    else:
        return redirect("/school/login")






