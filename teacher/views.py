from http.client import HTTPMessage
from django.http import Http404, HttpResponse, HttpResponseNotFound
from django.shortcuts import redirect, render
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from requests import session
from django.contrib.auth.decorators import login_required
from .models import teacher_profile

from school.models import School, Classroom, subjects
from django.utils.crypto import get_random_string
import openpyxl
from django.contrib import messages
# Create your views here.



# profile of the teacher
# Ayon
@login_required(login_url='/login')
def profile_teacher(request):
    if request.user.groups.filter(name='school').exists():
        # gets the teacher id of the teacher profile and checks if it exists
        if request.GET['teacher_id'] != '' and teacher_profile.objects.filter(id=request.GET['teacher_id'], school=request.user.schooluserprofile).exists():
            context = {
                'teacher' : teacher_profile.objects.get(id=request.GET['teacher_id'])
            }
            return render(request, 'teacher/profile.html', context)
        else:
            return HttpResponseNotFound('<div style="text-align:center;margin-top:400px;"><h2>Teacher not Found(HTTP_404)<h2><h4><a href="/school/">Home</a></h4></div>')
    else:
        return redirect('/school/school-logout/')


# get all the teachers
# Ayon
@login_required(login_url='/school/login')
def teachers(request):
    if request.user.groups.filter(name='school').exists():
        school = School.objects.get(user=User.objects.get(id=request.user.id))
        teachers = teacher_profile.objects.filter(school = school)
        return render(request, 'teacher/teachers.html' , {'teachers' : teachers, 'school' : school, 'schoolNavHeader' : True})
    return redirect('/school/school-logout/')


# login the user and set his usertype
# Ayon
def login_teacher(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            # User is authenticated

            # checks the user is Teacher or not
            if user.groups.filter(name='Teachers').exists():
                # if yes set the user type as Teachers and login

                # log in the teacher
                login(request, user)

                # starts a session teacher and and userType of the session is Teacher
                request.session['teacher'] = {
                'isLoggedIn' : True,
                'userType' : "Teachers"
                }
                messages.success(request, 'Login Successful')
                return redirect('/teacher?msg=login success')
            else :
                messages.error(request, 'Not a teacher')
                return redirect('/teacher/login?msg=invalid group')
        else:
            messages.error(request, 'Username and Passwords did not match!')
            return redirect("/teacher/login?msg=invalid credentials")
    return render(request, 'teacher/login.html')


# create the teacher account and teacher profile
# Ayon
@login_required(login_url='/school/school-login')
def create_teacher(request):
    # check if the school is logged in or not
    if request.user.groups.filter(name='school').exists():
        try:
            school = request.user.schooluserprofile
            if request.method == 'POST':
                first_name = request.POST['first_name']
                last_name = request.POST['last_name']
                email = request.POST['email']

                username = first_name+'_'+last_name
                while User.objects.filter(username=username).exists():
                    username = first_name+'_'+last_name+ '_' + get_random_string(4)
                password = get_random_string(8)

                # create a user account for the teacher
                teacher = User.objects.create_user(username, email, password)
                teacher.first_name = first_name
                teacher.last_name = last_name
                teacher.save()

                

                # add the teacher to the Teachers group
                Group.objects.get(name='Teachers').user_set.add(teacher)

                # create the teacher profile for the teacher account
                add_teacher_profile = teacher_profile.objects.create(
                    user = teacher, # the teacher account
                    school = school,
                    full_name  = f"{first_name} {last_name}",
                    phone = request.POST['phone'],
                    gender = request.POST['gender'],
                    subject = request.POST['subject'],
                    psw = password,
                )
                if len( request.FILES) != 0:
                    add_teacher_profile.img = request.FILES['img']
                classroom = request.POST.getlist('classroom')
                for i in classroom:
                    add_teacher_profile.classroom.add(Classroom.objects.get(id=i))
                add_teacher_profile.save()
                messages.success(request, 'Teacher created')
                return redirect(f'/teacher/profile?teacher_id={add_teacher_profile.id}&msg=add success')
            else:
                context = {
                    'school_id' : request.user.schooluserprofile.id,
                    'school' : school,
                    'schoolNavHeader' : True,
                    'classroom' : Classroom.objects.filter(school_id=request.user.schooluserprofile.id),
                    'subject' : subjects.objects.all(),
                }
                return render(request, 'teacher/add-teacher.html', context)
        except:
            messages.error(request, 'Add failed')
            return redirect(f'/teacher/add&msg=add failed')
    else:
        return redirect("/school/logout")


# update the teacher_profile
# Ayon
@login_required(login_url='/teacher/login')
def edit_teacher(request):
    if request.user.groups.filter(name='Teachers').exists():
        try:
            if request.method == 'POST':
                user = User.objects.get(id = request.user.id)
                user.first_name = request.POST['first_name']
                user.last_name = request.POST['last_name']
                user.save()


                teacher = teacher_profile.objects.filter(user=user).update(
                    full_name  = request.POST['first_name'] + ' ' + request.POST['last_name'],
                    phone = request.POST['phone'],
                    gender = request.POST['gender'],
                    subject = request.POST['subject'],
                )

                x = teacher_profile.objects.get(user=user)
                if len( request.FILES) != 0:
                    x.img = request.FILES['img']
                if  request.POST['address'] != '':
                    x.address = request.POST['address']
                if  request.POST['zipcode'] != '':
                    x.zipcode = request.POST['zipcode']

                classroom = request.POST.getlist('classroom')
                x.classroom.clear()
                for i in classroom:
                    x.classroom.add(Classroom.objects.get(id=i))    

                x.save()
                messages.success(request, 'Profile updated')
                return redirect('/teacher/update?msg=update success')
            else:
                user = User.objects.get(id = request.user.id)
                teacher = teacher_profile.objects.get(user=user)
                context = {
                    'teacher' : teacher,
                    'school_id' : teacher.school.id,
                    'classroom' : Classroom.objects.filter(school_id=teacher.school.id),
                    'subject' : subjects.objects.all(),
                }
                return render(request, 'teacher/edit-teacher.html', context)
        except:
            messages.error(request, 'Update Failed')
            return redirect('/teacher/update?msg=update failed')   
    return redirect('/teacher/logout')


# logout for the teacher
# Ayon
def logout_teacher(request):
    try :
        del request.session['teacher']
    except:
        pass
    messages.error(request, 'Logged out')
    logout(request)
    return redirect('/teacher/login')

# upload excel for the teacher
# Ayon
@login_required(login_url='/school/school-login')
def upload_teacher(request):
    if request.user.groups.filter(name='school').exists():
        try:
            school =  request.user.schooluserprofile
            if request.method == 'POST':

                if request.FILES['teacher_file']  != '':
                    data = openpyxl.load_workbook(request.FILES['teacher_file'])
                    data = data.active
                    data_list = []

                    for row in range(2, data.max_row):
                        for col in data.iter_cols(1, 7): #for col in data.iter_cols(1, data.max_column):
                            if (str(col[row].value) != '' and str(col[row].value) != None):
                                data_list.append(str(col[row].value))
                            else:
                                messages.error(request, f'Teachers Upload Failed (Check : cell no [{row}, {col}] #EMPTY_{str(col[row].value)})')
                                return redirect(f"/teacher/add/upload/?msg=upload failed (Check : cell no [{row}, {col}] #template error)")    

                        first_name = data_list[0]
                        last_name =  data_list[1]
                        email = data_list[2]
                        phone = data_list[3]

                        #GENDER FORMATTER
                        if(data_list[4].lower() == 'f'):
                            gender = "Female"
                        elif(data_list[4].lower() == 'm'):
                            gender = "Male"
                        elif(data_list[4].lower() == 'o'):
                            gender = "Others"
                        else:
                            messages.error(request, f'Teachers Upload Failed (Check : cell no [{row}, 5] GenderFormatError#Should be M/F/O : {data_list[4]})')
                            return redirect(f"/teacher/add/upload/?msg=upload failed (Check : cell no [{row}, 5] #GenderFormatError)") 

                        #CHECKS FOR VALID SUBJECT CODE 
                        if subjects.objects.filter(id=int(data_list[5])).exists():
                            subject = int(data_list[5])
                        else:
                            messages.error(request, f'Teachers Upload Failed (Check : cell no [{row}, 6] SubjectCodeError#Subject not exists.Create first : {data_list[5]})')
                            return redirect(f"/teacher/add/upload/?msg=upload failed (Check : cell no [{row}, 6] #SubjectCodeDoesNotExist)")   
        
                        try:
                            data_list[6].replace(' ', '')
                            classroom = list(data_list[6].split(','))
                        except:
                            messages.error(request, f'Teachers Upload Failed (Check : cell no [{row}, 7] ClassroomsFormatError#Should be as 2,5,7 [no_space]: {data_list[5]})')
                            return redirect(f"/teacher/add/upload/?msg=upload failed (Check : cell no [{row}, 7] #ClassroomsFormatError)") 
                        

                        username = first_name+'_'+last_name
                        while User.objects.filter(username=username).exists():
                            username = first_name+'_'+last_name+'_'+get_random_string(4)
                        password = get_random_string(8)

                        data_list.clear()
                        
                        full_name  = f"{first_name} {last_name}"
                        teacher = User.objects.create_user(username, email, password)
                        teacher.first_name = first_name
                        teacher.last_name = last_name
                        teacher.save()

                        # add the teacher to the Teachers group
                        Group.objects.get(name='Teachers').user_set.add(teacher)

                        # create the teacher profile for the teacher account
                        add_teacher_profile = teacher_profile.objects.create(
                            user = teacher, # the teacher account
                            school = school,
                            full_name  = full_name,
                            phone = phone,
                            gender = gender,
                            subject = subject,
                            psw = password
                        )
                        for i in classroom:
                            if Classroom.objects.filter(classroom_id=i, school_id=school).exists():
                                add_teacher_profile.classroom.add(Classroom.objects.get(classroom_id=i))
                            else:
                                messages.error(request, f"Teachers Upload Failed (Check : cell no [{row}, 7] ClassroomCodeError#Classroom can't be assigned: {i})")
                                return redirect(f"/teacher/add/upload/?msg=upload failed (Check : cell no [{row}, 7] #ClassroomCodeError)")   
                        add_teacher_profile.save()        
                        messages.success(request, 'Teachers Uploaded')
                    return redirect("/teacher/list")
                return redirect('/teacher/add/upload?id=')
            context = {
                'school' : school,
                'schoolNavHeader' : True
            }
            return render(request, 'teacher/add-upload.html', context)
        except:
            messages.error(request, 'Upload failed')
            return redirect('/teacher/add/upload?id=')
    return redirect('/school/school-logout')

# teacher dash
# Ayon
@login_required(login_url='/teacher/login')
def dash(request):
    if request.user.is_authenticated and teacher_profile.objects.filter(user=User.objects.get(id=request.user.id)).exists():
        teacher = teacher_profile.objects.get(id=request.user.teacherprofile.id)
        context = {
            'teacher' : teacher,
            'school_id' : teacher.school.id
        }
        return render(request, 'teacher/index.html', context)
    return redirect('/teacher/login')

    